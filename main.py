from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime
import os

with open('wetaca.html', 'r', encoding='utf-8') as file:
    html = file.read()
    soup = BeautifulSoup(html, 'html.parser')
name_elements = soup.find_all('div', class_='txt-action txt-action-s')
price_elements = soup.find_all('div', class_='txt-p-1 txt-p-1-s')

names = []
prices = []
for name_element in name_elements:
    names.append(name_element.text)

for price_element in price_elements:
    prices.append(price_element.text)

product_prices = dict(zip(names, prices))


wb = load_workbook('WetacaTemplate.xlsx')

ws = wb['Wetaca'] 

ws.auto_filter.ref = None
row_count = 4

for i, name in enumerate(names):
    if row_count > 46:
        break
    ws.cell(row=i+4, column=2).value = name
    price = float(prices[i].replace("€", "").replace(",","."))
    ws.cell(row=i+4, column=3).value = price
    row_count += 1

now = datetime.now()
date_string = now.strftime("%d%m%y")
my_docs = 'C:/Users/i81345/OneDrive - Verisk Analytics/Documents'
wb.save(os.path.join(my_docs, f"{date_string}_Wetaca.xlsx"))